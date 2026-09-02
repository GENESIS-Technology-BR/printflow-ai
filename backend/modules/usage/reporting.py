from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Iterable

from backend.modules.printers.model import Printer

from .model import PrinterUsageDaily


BRAND_NAVY = "0B2A52"
BRAND_BLUE = "0A6ED1"
BRAND_TEAL = "25C6B7"
BRAND_GREEN = "21D89B"
BRAND_MUTED = "6B7C93"


def _looks_like_opaque_hex(value: str | None) -> bool:
    if not value:
        return False
    text = value.strip().lower()
    if text.startswith("0x"):
        text = text[2:]
    return len(text) >= 48 and all(char in "0123456789abcdef" for char in text)


def _clean_identity(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if not text or _looks_like_opaque_hex(text):
        return None
    return text


def _display_name(
    custom_name: str | None,
    hostname: str | None,
    name: str | None,
    ip: str | None,
) -> str:
    if custom_name and custom_name.strip():
        return custom_name.strip()

    for candidate in (hostname, name):
        cleaned = _clean_identity(candidate)
        if cleaned:
            return cleaned

    return ip or "Impressora sem nome"


def _format_number(value: int | None) -> str:
    if value is None:
        return "-"
    return f"{int(value):,}".replace(",", ".")


def _format_currency(value: float) -> str:
    text = f"{float(value):,.2f}"
    text = text.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {text}"


def _format_rate(value: float) -> str:
    return f"R$ {float(value):.4f}".replace(".", ",")


def _cost_value(value) -> float:
    if value is None:
        return 0.0
    try:
        return max(float(value), 0.0)
    except (TypeError, ValueError):
        return 0.0


def _model_label(manufacturer: str | None, model: str | None) -> str:
    brand = (manufacturer or "").strip()
    model_text = (model or "").strip()

    if model_text:
        if brand and model_text.lower().startswith(brand.lower()):
            return model_text
        return " ".join(part for part in (brand, model_text) if part)

    return brand or "-"


def _build_brand_icon_png() -> BytesIO:
    from PIL import Image, ImageDraw

    image = Image.new("RGBA", (180, 180), (255, 255, 255, 0))
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((58, 18, 164, 162), radius=24, fill="#0A6ED1")

    circuit_rows = (
        (62, "#1C86F3"),
        (90, "#25C6B7"),
        (118, "#21D89B"),
    )
    for y, color in circuit_rows:
        draw.line((18, y, 82, y), fill=color, width=10)
        draw.ellipse((7, y - 11, 29, y + 11), fill=color)
        draw.ellipse((71, y - 11, 93, y + 11), fill=color)

    stream = BytesIO()
    image.save(stream, format="PNG")
    stream.seek(0)
    return stream


def consolidate_usage(
    history: Iterable[PrinterUsageDaily],
    printers: Iterable[Printer] = (),
    default_cost_per_page=0,
) -> list[dict]:
    printer_list = list(printers)
    printer_map = {printer.uuid: printer for printer in printer_list}
    groups: dict[str, dict] = {}
    ordered_history = sorted(
        history,
        key=lambda row: (
            row.printer_uuid,
            row.usage_date,
            row.id or 0,
        ),
    )

    for usage in ordered_history:
        item = groups.get(usage.printer_uuid)
        if item is None:
            item = {
                "printer_uuid": usage.printer_uuid,
                "display_name": _display_name(
                    usage.custom_name,
                    usage.hostname,
                    usage.name,
                    usage.ip,
                ),
                "ip": usage.ip,
                "hostname": _clean_identity(usage.hostname),
                "manufacturer": usage.manufacturer,
                "model": usage.model,
                "serial": _clean_identity(usage.serial),
                "unit_name": usage.unit_name,
                "sector_name": usage.sector_name,
                "first_usage_date": usage.usage_date,
                "last_usage_date": usage.usage_date,
                "opening_page_count": usage.opening_page_count,
                "closing_page_count": usage.closing_page_count,
                "pages_printed": 0,
                "anomaly_count": 0,
                "last_anomaly_type": None,
            }
            groups[usage.printer_uuid] = item

        item["pages_printed"] += usage.pages_printed
        item["anomaly_count"] += usage.anomaly_count
        item["last_usage_date"] = usage.usage_date
        item["closing_page_count"] = usage.closing_page_count
        item["display_name"] = _display_name(
            usage.custom_name,
            usage.hostname,
            usage.name,
            usage.ip,
        )
        item["ip"] = usage.ip
        item["hostname"] = _clean_identity(usage.hostname)
        item["manufacturer"] = usage.manufacturer
        item["model"] = usage.model
        item["serial"] = _clean_identity(usage.serial)
        item["unit_name"] = usage.unit_name
        item["sector_name"] = usage.sector_name

        if usage.last_anomaly_type:
            item["last_anomaly_type"] = usage.last_anomaly_type

    for printer in printer_list:
        if printer.uuid in groups:
            continue

        page_count = int(printer.page_count) if printer.page_count is not None else None
        groups[printer.uuid] = {
            "printer_uuid": printer.uuid,
            "display_name": _display_name(
                printer.custom_name,
                printer.hostname,
                printer.name,
                printer.ip,
            ),
            "ip": printer.ip,
            "hostname": _clean_identity(printer.hostname),
            "manufacturer": printer.manufacturer,
            "model": printer.model,
            "serial": _clean_identity(printer.serial),
            "unit_name": printer.unit_name,
            "sector_name": printer.sector_name,
            "first_usage_date": None,
            "last_usage_date": None,
            "opening_page_count": page_count,
            "closing_page_count": page_count,
            "pages_printed": 0,
            "anomaly_count": 0,
            "last_anomaly_type": None,
        }

    default_cost = _cost_value(default_cost_per_page)
    for printer_uuid, item in groups.items():
        printer = printer_map.get(printer_uuid)
        override = getattr(printer, "cost_per_page", None) if printer else None
        effective_cost = _cost_value(override) if override is not None else default_cost
        item["cost_per_page"] = round(effective_cost, 4)
        item["estimated_cost"] = round(item["pages_printed"] * effective_cost, 2)
        item["cost_source"] = "printer" if override is not None else "company"

    return sorted(
        groups.values(),
        key=lambda item: (
            item["unit_name"] or "",
            item["sector_name"] or "",
            item["display_name"].lower(),
        ),
    )


def build_excel_report(
    company_name: str,
    start: date,
    end: date,
    rows: list[dict],
    history: Iterable[PrinterUsageDaily],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumo"
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "Printflow - Relatorio de Impressao"
    sheet["A1"].font = Font(size=18, bold=True, color=BRAND_NAVY)
    sheet["A2"] = f"Empresa: {company_name}"
    sheet["A2"].font = Font(size=10, color=BRAND_MUTED)
    sheet["A3"] = (
        f"Periodo: {start.strftime('%d/%m/%Y')} "
        f"a {end.strftime('%d/%m/%Y')}"
    )
    sheet["A3"].font = Font(size=10, color=BRAND_MUTED)

    logo_stream = _build_brand_icon_png()
    logo = XLImage(logo_stream)
    logo.width = 48
    logo.height = 48
    sheet.add_image(logo, "N1")

    headers = [
        "Impressora", "IP", "Hostname", "Fabricante", "Modelo", "Serial",
        "Unidade", "Setor", "Primeira leitura", "Ultima leitura",
        "Contador inicial", "Contador final", "Impressoes no periodo", "Anomalias",
        "Custo/pagina (R$)", "Custo estimado (R$)",
    ]
    header_row = 5

    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        cell.alignment = Alignment(horizontal="center")

    for row_index, item in enumerate(rows, start=header_row + 1):
        values = [
            item["display_name"], item["ip"] or "", item["hostname"] or "",
            item["manufacturer"] or "", item["model"] or "", item["serial"] or "",
            item["unit_name"] or "", item["sector_name"] or "",
            item["first_usage_date"].strftime("%d/%m/%Y") if item["first_usage_date"] else "Sem historico",
            item["last_usage_date"].strftime("%d/%m/%Y") if item["last_usage_date"] else "Sem historico",
            item["opening_page_count"], item["closing_page_count"],
            item["pages_printed"], item["anomaly_count"],
            item["cost_per_page"], item["estimated_cost"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            if column == 15:
                cell.number_format = 'R$ #,##0.0000'
            elif column == 16:
                cell.number_format = 'R$ #,##0.00'

    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:P{max(header_row, header_row + len(rows))}"
    widths = [30, 16, 24, 18, 28, 22, 20, 20, 16, 16, 16, 16, 20, 12, 18, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    detail = workbook.create_sheet("Historico diario")
    detail.sheet_view.showGridLines = False
    detail_headers = [
        "Data", "Impressora", "IP", "Unidade", "Setor",
        "Contador abertura", "Contador fechamento", "Impressoes", "Anomalias",
    ]
    for column, label in enumerate(detail_headers, start=1):
        cell = detail.cell(row=1, column=column, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)

    for row_index, usage in enumerate(
        sorted(history, key=lambda item: (item.usage_date, item.printer_uuid)),
        start=2,
    ):
        values = [
            usage.usage_date.strftime("%d/%m/%Y"),
            _display_name(usage.custom_name, usage.hostname, usage.name, usage.ip),
            usage.ip, usage.unit_name or "", usage.sector_name or "",
            usage.opening_page_count, usage.closing_page_count,
            usage.pages_printed, usage.anomaly_count,
        ]
        for column, value in enumerate(values, start=1):
            detail.cell(row=row_index, column=column, value=value)

    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:I{max(1, detail.max_row)}"
    for index, width in enumerate([14, 30, 16, 20, 20, 18, 18, 14, 12], start=1):
        detail.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf_report(
    company_name: str,
    start: date,
    end: date,
    rows: list[dict],
) -> bytes:
    from reportlab.graphics.shapes import Circle, Drawing, Line, Rect
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=15 * mm,
        title="Printflow - Relatorio de Impressao",
    )
    styles = getSampleStyleSheet()

    mark = Drawing(17 * mm, 17 * mm)
    mark.add(Rect(6 * mm, 1 * mm, 10 * mm, 15 * mm, rx=2 * mm, ry=2 * mm,
                  fillColor=colors.HexColor(f"#{BRAND_BLUE}"), strokeColor=None))
    for y, color in ((12 * mm, "#1C86F3"), (8.5 * mm, "#25C6B7"), (5 * mm, "#21D89B")):
        mark.add(Line(1.5 * mm, y, 8.2 * mm, y, strokeColor=colors.HexColor(color), strokeWidth=1.5 * mm))
        mark.add(Circle(1.5 * mm, y, 1.4 * mm, fillColor=colors.HexColor(color), strokeColor=None))
        mark.add(Circle(8.2 * mm, y, 1.4 * mm, fillColor=colors.HexColor(color), strokeColor=None))

    brand_copy = Paragraph(
        (
            f'<font color="#{BRAND_NAVY}" size="18"><b>Printflow</b></font><br/>'
            f'<font color="#{BRAND_MUTED}" size="8">Gestão inteligente de impressão</font>'
        ),
        styles["Normal"],
    )
    brand_header = Table([[mark, brand_copy]], colWidths=[20 * mm, 82 * mm])
    brand_header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    story = [
        brand_header,
        Spacer(1, 2.5 * mm),
        Paragraph("Relatorio de Impressao", styles["Title"]),
        Paragraph(f"<b>Empresa:</b> {company_name}", styles["Normal"]),
        Paragraph(
            f"<b>Periodo:</b> {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}",
            styles["Normal"],
        ),
        Paragraph(
            "<b>Custos:</b> estimativa calculada pelas tarifas atualmente configuradas.",
            styles["Normal"],
        ),
        Spacer(1, 5 * mm),
    ]

    table_data = [[
        "Impressora", "IP", "Unidade / Setor", "Modelo",
        "Inicial", "Final", "Impressoes", "R$/pag.", "Custo estimado",
    ]]
    for item in rows:
        organization = " / ".join(
            part for part in (item["unit_name"], item["sector_name"]) if part
        ) or "-"
        model = _model_label(item["manufacturer"], item["model"])
        table_data.append([
            item["display_name"],
            item["ip"] or "-",
            organization,
            model,
            _format_number(item["opening_page_count"]),
            _format_number(item["closing_page_count"]),
            _format_number(item["pages_printed"]),
            _format_rate(item["cost_per_page"]),
            _format_currency(item["estimated_cost"]),
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[42 * mm, 23 * mm, 38 * mm, 48 * mm, 22 * mm, 22 * mm, 24 * mm, 24 * mm, 28 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_BLUE}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.6),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#A8B8C7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F8FC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    total_pages = sum(item["pages_printed"] for item in rows)
    total_cost = sum(float(item["estimated_cost"]) for item in rows)
    story.extend([
        Spacer(1, 5 * mm),
        Paragraph(
            (
                f"<b>Total de impressoras:</b> {len(rows)} &nbsp;&nbsp; "
                f"<b>Total de impressoes:</b> {_format_number(total_pages)} &nbsp;&nbsp; "
                f"<b>Custo estimado:</b> {_format_currency(total_cost)}"
            ),
            styles["Normal"],
        ),
    ])

    def draw_footer(canvas, _document) -> None:
        canvas.saveState()
        width, _height = page_size
        canvas.setStrokeColor(colors.HexColor("#D7E1EA"))
        canvas.setLineWidth(0.4)
        canvas.line(10 * mm, 9 * mm, width - 10 * mm, 9 * mm)
        canvas.setFillColor(colors.HexColor(f"#{BRAND_MUTED}"))
        canvas.setFont("Helvetica", 7)
        canvas.drawString(10 * mm, 5 * mm, f"Printflow · {company_name}")
        canvas.drawRightString(
            width - 10 * mm,
            5 * mm,
            f"Pagina {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    document.build(
        story,
        onFirstPage=draw_footer,
        onLaterPages=draw_footer,
    )
    return output.getvalue()
