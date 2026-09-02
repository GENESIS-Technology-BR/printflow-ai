from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from backend.modules.usage.reporting import build_excel_report, build_pdf_report


SAMPLE_ROWS = [
    {
        "printer_uuid": "printer-001",
        "display_name": "Impressora Financeiro",
        "ip": "10.0.0.25",
        "hostname": "PRN-FIN-01",
        "manufacturer": "HP",
        "model": "LaserJet",
        "serial": "SN001",
        "unit_name": "Matriz",
        "sector_name": "Financeiro",
        "first_usage_date": date(2026, 9, 1),
        "last_usage_date": date(2026, 9, 2),
        "opening_page_count": 1000,
        "closing_page_count": 1125,
        "pages_printed": 125,
        "anomaly_count": 0,
        "last_anomaly_type": None,
        "cost_per_page": 0.12,
        "estimated_cost": 15.0,
        "cost_source": "company",
    }
]


def test_excel_report_generates_valid_workbook():
    content = build_excel_report(
        "Empresa Teste",
        date(2026, 9, 1),
        date(2026, 9, 2),
        SAMPLE_ROWS,
        [],
    )

    assert content.startswith(b"PK")

    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames == ["Resumo", "Historico diario"]

    summary = workbook["Resumo"]
    assert summary["A1"].value == "PRINTFLOW AI - Relatorio de Impressao"
    assert summary["A2"].value == "Empresa: Empresa Teste"
    assert summary["A6"].value == "Impressora Financeiro"
    assert summary["M6"].value == 125
    assert summary["O6"].value == 0.12
    assert summary["P6"].value == 15.0


def test_pdf_report_generates_valid_pdf():
    content = build_pdf_report(
        "Empresa Teste",
        date(2026, 9, 1),
        date(2026, 9, 2),
        SAMPLE_ROWS,
    )

    assert content.startswith(b"%PDF")
    assert len(content) > 1000
